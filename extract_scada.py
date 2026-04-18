"""
SCADA Mnemonic Extractor
Reads numeric values from mnemonic screen videos and fills Excel tables
per frame (or per unique second).

Usage:
    python extract_scada.py --video_idx 1   # process Видео 1 → Таблица 1
    python extract_scada.py --all           # process all three videos
"""

import os
import sys
import re
import argparse
import time
import warnings
import logging
from pathlib import Path

import cv2
import numpy as np
from PIL import Image
import openpyxl

# Suppress warnings
warnings.filterwarnings("ignore")
logging.getLogger("easyocr").setLevel(logging.ERROR)

# ─── Paths ─────────────────────────────────────────────────────────────────
BASE = Path(r"C:\Users\kudzh_o\Documents\GitHub\hac\Хакатон «ИИ – АВТОМАТИЗАЦИЯ»\Задание 2")
OUTPUT_DIR = Path(r"C:\Users\kudzh_o\Documents\GitHub\hac\results")
OUTPUT_DIR.mkdir(exist_ok=True)


def load_easyocr():
    import easyocr
    reader = easyocr.Reader(['ru', 'en'], gpu=False, verbose=False)
    return reader


class OCRAdapter:
    """Adapter to provide unified `readtext(frame)` across OCR backends.

    Supported backends: 'easyocr' (default), 'paddlevl' (PaddleOCR variant).
    """
    def __init__(self, backend='easyocr', det_model_dir=None, rec_model_dir=None):
        self.backend = backend
        self._impl = None
        self._det_model_dir = det_model_dir
        self._rec_model_dir = rec_model_dir
        if backend == 'easyocr':
            try:
                import easyocr
                self._impl = easyocr.Reader(['ru', 'en'], gpu=False, verbose=False)
            except Exception as e:
                raise RuntimeError(f"Failed to init EasyOCR: {e}")
        elif backend in ('paddlevl', 'paddle'):
            try:
                from paddleocr import PaddleOCR
                # Use a generic PaddleOCR initialization; allow specifying model dirs
                kwargs = {'use_angle_cls': False, 'lang': 'ru'}
                if self._det_model_dir:
                    kwargs['det_model_dir'] = str(self._det_model_dir)
                if self._rec_model_dir:
                    kwargs['rec_model_dir'] = str(self._rec_model_dir)
                self._impl = PaddleOCR(**kwargs)
            except Exception as e:
                raise RuntimeError(f"Failed to init PaddleOCR: {e}")
        elif backend in ('tesseract', 'pytesseract'):
            try:
                import pytesseract
                from pytesseract import Output
                from pathlib import Path
                # Allow user to override tesseract binary via environment var
                tcmd = os.environ.get('TESSERACT_CMD')
                if not tcmd:
                    # Try to find tesseract automatically
                    candidates = [
                        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                        os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
                        r"C:\tesseract-5.5.2\tesseract.exe",
                        r"D:\tesseract-5.5.2\tesseract.exe",
                        r"E:\tesseract-5.5.2\tesseract.exe",
                    ]
                    for candidate in candidates:
                        if Path(candidate).exists():
                            tcmd = candidate
                            print(f"Found Tesseract at: {tcmd}")
                            break
                if tcmd:
                    pytesseract.pytesseract.tesseract_cmd = tcmd
                self._impl = pytesseract
                self._tcmd = tcmd or 'tesseract'
                self._tesseract_output = Output
                # Check if Russian language data is available
                try:
                    import subprocess
                    result = subprocess.run(
                        [self._tcmd, '--list-langs'],
                        capture_output=True, text=True, timeout=10,
                        creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if hasattr(subprocess, 'CREATE_NEW_PROCESS_GROUP') else 0,
                    )
                    self._has_rus_lang = 'rus' in result.stdout
                except Exception:
                    self._has_rus_lang = False
            except Exception as e:
                raise RuntimeError(f"Failed to init pytesseract: {e}")
        else:
            raise ValueError(f"Unknown OCR backend: {backend}")

    def readtext(self, frame_rgb, detail=1, paragraph=False, allowlist=None):
        """Return list of (bbox, text, conf) similar to EasyOCR format."""
        if self.backend == 'easyocr':
            return self._impl.readtext(frame_rgb, detail=1, paragraph=False)

        if self.backend in ('tesseract', 'pytesseract'):
            # Call tesseract directly via subprocess to avoid Python 3.12+Windows
            # KeyboardInterrupt bug in pytesseract's subprocess.communicate()
            return self._tesseract_readtext_direct(frame_rgb)

        # PaddleOCR: convert output to (bbox, text, conf)
        # Different PaddleOCR versions accept different kwargs; call without 'cls'
        try:
            results = self._impl.ocr(frame_rgb)
        except TypeError:
            results = self._impl.ocr(frame_rgb, cls=False)
        out = []
        if not results:
            return out

        def extract_from_item(item):
            # item can be (box, (text, score)) or (box, text, score) or nested lists
            if not item:
                return None
            # If it's a 3-tuple like (box, text, score)
            if isinstance(item, (list, tuple)) and len(item) == 3 and isinstance(item[0], (list, tuple)):
                box = item[0]
                text = item[1]
                try:
                    conf = float(item[2])
                except Exception:
                    conf = 0.0
                return (box, str(text), conf)
            # If it's (box, (text, score))
            if isinstance(item, (list, tuple)) and len(item) >= 2:
                box = item[0]
                txt_score = item[1]
                if isinstance(txt_score, (list, tuple)) and len(txt_score) >= 2:
                    try:
                        confv = float(txt_score[1])
                    except Exception:
                        confv = 0.0
                    return (box, str(txt_score[0]), confv)
                else:
                    return (box, str(txt_score), 0.0)
            return None

        # Results may be nested: either list of lines or list of pages
        for entry in results:
            # If entry itself looks like a single detection
            det = extract_from_item(entry)
            if det:
                out.append(det)
                continue
            # Otherwise, try to iterate sub-items
            if isinstance(entry, (list, tuple)):
                for sub in entry:
                    det = extract_from_item(sub)
                    if det:
                        out.append(det)
        return out

    def _tesseract_readtext_direct(self, frame_rgb):
        """Call tesseract via subprocess directly to avoid pytesseract's
        subprocess.communicate() KeyboardInterrupt bug on Python 3.12 + Windows."""
        import subprocess
        import tempfile
        import csv
        tcmd = self._tcmd
        try:
            img = Image.fromarray(frame_rgb)
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_in:
                tmp_in_path = tmp_in.name
                img.save(tmp_in_path)
            with tempfile.NamedTemporaryFile(suffix='', delete=False) as tmp_out:
                tmp_out_path = tmp_out.name
            # Determine available langs
            langs = 'rus+eng' if self._has_rus_lang else 'eng'
            proc = subprocess.Popen(
                [tcmd, tmp_in_path, tmp_out_path, '--oem', '3', '--psm', '11',
                 '-l', langs, 'tsv'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if hasattr(subprocess, 'CREATE_NEW_PROCESS_GROUP') else 0,
            )
            try:
                proc.wait(timeout=30)
            except subprocess.TimeoutExpired:
                proc.kill()
                proc.wait()
                return []
            except KeyboardInterrupt:
                proc.kill()
                try:
                    proc.wait()
                except Exception:
                    pass
                return []
            tsv_path = tmp_out_path + '.tsv'
            out = []
            try:
                import os as _os
                if _os.path.exists(tsv_path):
                    with open(tsv_path, encoding='utf-8', newline='') as f:
                        reader_csv = csv.DictReader(f, delimiter='\t')
                        for row in reader_csv:
                            txt = str(row.get('text', '')).strip()
                            if not txt:
                                continue
                            try:
                                conf = float(row.get('conf', 0)) / 100.0
                            except Exception:
                                conf = 0.0
                            if conf < 0:
                                continue
                            try:
                                x = int(row.get('left', 0))
                                y = int(row.get('top', 0))
                                w = int(row.get('width', 0))
                                h = int(row.get('height', 0))
                            except Exception:
                                continue
                            box = [[x, y], [x + w, y], [x + w, y + h], [x, y + h]]
                            out.append((box, txt, conf))
            finally:
                import os as _os
                for p in [tmp_in_path, tmp_out_path, tsv_path]:
                    try:
                        _os.unlink(p)
                    except Exception:
                        pass
            return out
        except (Exception, KeyboardInterrupt):
            return []


def read_video_frame(cap, frame_idx):
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
    ret, frame = cap.read()
    if not ret:
        return None
    return cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)


def pil_save(frame_rgb, path):
    """Save numpy RGB array via PIL (avoids cv2 Cyrillic path issue)."""
    Image.fromarray(frame_rgb).save(path)


def ocr_frame(reader, frame_rgb):
    """Run EasyOCR on a full frame. Returns list of (bbox, text, conf)."""
    max_dim = max(frame_rgb.shape[:2]) if getattr(frame_rgb, 'shape', None) is not None else 0
    if max_dim > 1400:
        scale = 1400.0 / float(max_dim)
        resized = cv2.resize(frame_rgb, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
        result = reader.readtext(resized, detail=1, paragraph=False)
        scaled_result = []
        inv_scale = 1.0 / scale
        for item in result:
            if not isinstance(item, (list, tuple)) or len(item) < 3:
                continue
            bbox, text, conf = item[0], item[1], item[2]
            try:
                scaled_bbox = [[float(x) * inv_scale, float(y) * inv_scale] for x, y in bbox]
            except Exception:
                scaled_bbox = bbox
            scaled_result.append((scaled_bbox, text, conf))
        return scaled_result

    result = reader.readtext(frame_rgb, detail=1, paragraph=False)
    return result


def is_number(text):
    """Check if a string looks like a numeric value."""
    t = text.strip().replace(',', '.').replace('−', '-').replace('–', '-')
    try:
        float(t)
        return True
    except ValueError:
        return False


def parse_number(text):
    """Parse text to float; returns None if not parseable."""
    t = text.strip().replace(',', '.').replace('−', '-').replace('–', '-').replace(' ', '')
    # Remove trailing non-numeric garbage
    m = re.match(r'^-?[\d]+\.?[\d]*', t)
    if m:
        try:
            return float(m.group())
        except Exception:
            pass
    return None


def bbox_center(bbox):
    """Get center of a polygon bbox returned by EasyOCR."""
    pts = np.array(bbox)
    # Handle empty/invalid bbox gracefully
    if pts.size == 0:
        return np.array([0.0, 0.0])
    try:
        return pts.mean(axis=0)
    except Exception:
        return np.array([0.0, 0.0])


def build_number_map(ocr_results):
    """
    From OCR results, return dict: (cx, cy) -> float value
    for every text element that is a number.
    Only keeps numeric items.
    """
    number_map = {}
    for bbox, text, conf in ocr_results:
        val = parse_number(text)
        if val is not None and conf > 0.3:
            cx, cy = bbox_center(bbox)
            number_map[(float(cx), float(cy))] = (val, conf, text)
    return number_map


def load_table_params(table_path):
    """
    Load parameter list from Excel table.
    Returns list of dicts: {name, unit, short, decimals}
    """
    wb = openpyxl.load_workbook(table_path)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active worksheet in {table_path}")
    params = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name is None:
            continue
        if isinstance(name, str) and name.strip():
            unit = row[1] if len(row) > 1 else ''
            short = row[2] if len(row) > 2 else ''
            decimals = row[3] if len(row) > 3 else 1
            try:
                decimals_value = int(str(decimals).strip())
            except Exception:
                decimals_value = 1
            params.append({
                'name': str(name).strip(),
                'unit': str(unit).strip() if unit else '',
                'short': str(short).strip() if short else '',
                'decimals': decimals_value,
            })
    return params


def detect_scene_change(prev_gray, curr_gray, threshold=3.0):
    """Detect if the frame content changed significantly (new time step)."""
    if prev_gray is None:
        return True
    diff = cv2.absdiff(prev_gray, curr_gray)
    return diff.mean() > threshold


def extract_all_numbers_from_frame(reader, frame_rgb):
    """
    Returns a flat list of numbers found in this frame with their positions.
    [(x, y, value, text, conf), ...]
    """
    results = ocr_frame(reader, frame_rgb)
    numbers = []
    for bbox, text, conf in results:
        val = parse_number(text)
        if val is not None and conf > 0.3:
            cx, cy = bbox_center(bbox)
            numbers.append((float(cx), float(cy), val, text, conf))
    return numbers


def match_numbers_to_params_by_position(numbers, regions):
    """
    Given detected numbers and known parameter regions (from layout detection),
    match each number to the nearest region.
    regions: list of (name, cx, cy)
    Returns dict: param_name -> value
    """
    result = {}
    for name, rx, ry in regions:
        # Find closest number to this region center
        best_val = None
        best_dist = float('inf')
        for nx, ny, val, text, conf in numbers:
            dist = ((nx - rx) ** 2 + (ny - ry) ** 2) ** 0.5
            if dist < best_dist:
                best_dist = dist
                best_val = val
        if best_dist < 80:  # max 80px away
            result[name] = best_val
    return result


def auto_detect_layout(reader, frame_rgb, params):
    """
    On first frame, run full OCR to find where each parameter's numeric value is.
    Strategy: look for short labels (T, P, dP, N, etc.) and grab the nearby number.
    
    Returns list of (param_name, value_cx, value_cy) — the screen positions.
    """
    results = ocr_frame(reader, frame_rgb)
    
    # Separate text items and numeric items
    text_items = []
    num_items = []
    for bbox, text, conf in results:
        val = parse_number(text)
        cx, cy = bbox_center(bbox)
        if val is not None and conf > 0.3:
            num_items.append((cx, cy, val, text, conf, bbox))
        elif conf > 0.4:
            text_items.append((cx, cy, text, conf, bbox))
    
    print(f"  Layout detection: {len(text_items)} text labels, {len(num_items)} numbers on screen")
    return text_items, num_items


def label_regions_from_text(regions, text_items, params, label_radius=120):
    """
    Try to assign a label to each numeric region by looking at nearby text items.
    
    For each number region, find the closest text label within label_radius px.
    Then try to match that text to a param's short name.
    
    Returns the regions list with added 'label' and 'param_name' fields.
    """
    # Build short-name lookup from params
    short_names = {}
    for p in params:
        sn = p.get('short', '').strip().lower()
        if sn and sn not in ('r', ''):
            if sn not in short_names:
                short_names[sn] = []
            short_names[sn].append(p['name'])
    
    labeled = []
    used_param_counts = {}  # track how many times each short name used
    
    for reg in regions:
        rx, ry = reg['cx'], reg['cy']
        
        # Find closest text item within radius
        best_text = None
        best_dist = float('inf')
        for tx, ty, text, conf, bbox in text_items:
            dist = ((tx - rx) ** 2 + (ty - ry) ** 2) ** 0.5
            if dist < label_radius and dist < best_dist:
                best_dist = dist
                best_text = text.strip()
        
        if best_text:
            # Try to match to a param short name
            best_text_lower = best_text.lower().strip()
            
            # Exact match
            if best_text_lower in short_names:
                cnt = used_param_counts.get(best_text_lower, 0)
                params_list = short_names[best_text_lower]
                param_name = params_list[min(cnt, len(params_list) - 1)]
                used_param_counts[best_text_lower] = cnt + 1
                reg = dict(reg, label=best_text, param_name=param_name)
            else:
                reg = dict(reg, label=best_text, param_name=None)
        else:
            reg = dict(reg, label=None, param_name=None)
        
        labeled.append(reg)
    
    matched = sum(1 for r in labeled if r.get('param_name'))
    print(f"  Region labeling: {matched}/{len(labeled)} regions matched to parameters")
    return labeled


def process_video(video_idx, frame_step=None, max_frames=None, reader=None):
    """
    Process one video and fill the corresponding Excel table.
    
    frame_step: how many frames to skip between readings (default: ~1 per second)
    """
    video_dir = BASE / f"Видео {video_idx}"
    video_path = video_dir / f"Видео {video_idx}.mp4"
    table_path = video_dir / f"Таблица {video_idx}.xlsx"
    
    print(f"\n{'='*60}")
    print(f"Processing: Видео {video_idx}")
    print(f"  Video: {video_path}")
    print(f"  Table: {table_path}")
    
    # Load parameters from table
    params = load_table_params(table_path)
    print(f"  Parameters in table: {len(params)}")
    
    # Open video
    cap = cv2.VideoCapture(str(video_path))
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    duration = total_frames / fps if fps > 0 else 0
    
    print(f"  Video: {width}x{height}, {fps:.1f} fps, {total_frames} frames, {duration:.1f}s")
    
    if frame_step is None:
        frame_step = max(1, int(fps))  # 1 frame per second
    
    print(f"  Frame step: {frame_step} (sampling every {frame_step/fps:.2f}s)")
    
    # Initialize OCR (use provided reader or default to EasyOCR adapter)
    print("  Loading OCR model...")
    if reader is None:
        reader = OCRAdapter('easyocr')
    print(f"  OCR ready ({getattr(reader, 'backend', 'easyocr')}).")
    
    # Collect data: list of (frame_idx, timestamp, {param_name: value})
    all_rows = []
    prev_gray = None
    frame_idx = 0
    processed = 0
    
    print(f"  Processing frames...")
    
    while True:
        if max_frames and processed >= max_frames:
            break
        
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ret, bgr = cap.read()
        if not ret:
            break
        
        gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
        timestamp = frame_idx / fps
        
        # Always process this frame (we step by frame_step)
        frame_rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        
        # Extract all numbers from frame
        numbers = extract_all_numbers_from_frame(reader, frame_rgb)
        
        if numbers:
            row_data = {
                '_frame': frame_idx,
                '_timestamp': f"{int(timestamp//60):02d}:{int(timestamp%60):02d}.{int((timestamp%1)*10):01d}",
                '_num_values': len(numbers),
            }
            # Store all detected numbers sorted by position (top-to-bottom, left-to-right)
            numbers_sorted = sorted(numbers, key=lambda x: (int(x[1]/30), int(x[0]/30)))
            for i, (nx, ny, val, text, conf) in enumerate(numbers_sorted):
                row_data[f'value_{i+1}'] = val
                row_data[f'pos_{i+1}'] = f"({int(nx)},{int(ny)})"
        else:
            row_data = {
                '_frame': frame_idx,
                '_timestamp': f"{int(timestamp//60):02d}:{int(timestamp%60):02d}.{int((timestamp%1)*10):01d}",
                '_num_values': 0,
            }
        
        all_rows.append(row_data)
        prev_gray = gray
        processed += 1
        
        if processed % 10 == 0:
            print(f"    Processed {processed} frames (frame {frame_idx}/{total_frames}, t={timestamp:.1f}s), found {len(numbers)} numbers")
        
        frame_idx += frame_step
    
    cap.release()
    print(f"  Total frames processed: {processed}")
    
    # Save results to Excel
    output_path = OUTPUT_DIR / f"Результаты_{video_idx}.xlsx"
    save_results_to_excel(all_rows, params, output_path, video_idx)
    print(f"  Results saved: {output_path}")
    
    return all_rows


def save_results_to_excel(all_rows, params, output_path, video_idx):
    """
    Save extracted data to Excel.
    Each row = one frame.
    Columns: Timestamp + all detected values (in order of position).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = f"Данные видео {video_idx}"
    
    # Find max number of values in any row
    max_vals = max((row.get('_num_values', 0) for row in all_rows), default=0)
    
    # Header row
    headers = ['Кадр', 'Временная метка', 'Кол-во значений']
    for i in range(1, max_vals + 1):
        headers.append(f'Значение {i}')
        headers.append(f'Позиция {i}')
    
    ws.append(headers)
    
    # Style header
    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row in all_rows:
        data = [
            row.get('_frame', ''),
            row.get('_timestamp', ''),
            row.get('_num_values', 0),
        ]
        for i in range(1, max_vals + 1):
            data.append(row.get(f'value_{i}', ''))
            data.append(row.get(f'pos_{i}', ''))
        ws.append(data)
    
    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        first_cell = col[0]
        column_letter = getattr(first_cell, 'column_letter', None)
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_len + 2, 30)
    
    # Also create a second sheet with parameter list for reference
    ws2 = wb.create_sheet("Параметры таблицы")
    ws2.append(['#', 'Наименование параметра', 'Единица', 'Короткое название', 'Знаков'])
    for i, p in enumerate(params, 1):
        ws2.append([i, p['name'], p['unit'], p['short'], p['decimals']])
    
    wb.save(output_path)


def process_video_smart(video_idx, frame_step=None, max_frames=None, reader=None):
    """
    Smarter version: 
    1. Run full OCR on first frame to build a position-to-parameter map.
    2. For each subsequent frame, only crop small regions around known positions.
    3. Run fast digit-only OCR on each cropped region.
    
    This gives per-frame rows with named columns matching table parameters.
    reader: optional pre-loaded EasyOCR reader (if None, loads a new one)
    """
    video_dir = BASE / f"Видео {video_idx}"
    video_path = video_dir / f"Видео {video_idx}.mp4"
    table_path = video_dir / f"Таблица {video_idx}.xlsx"
    
    print(f"\n{'='*60}")
    print(f"[Smart] Processing: Видео {video_idx}")
    
    params = load_table_params(table_path)
    print(f"  Parameters defined in table: {len(params)}")
    
    cap = cv2.VideoCapture(str(video_path))
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    duration = total_frames / fps if fps > 0 else 0
    
    print(f"  {width}x{height} @ {fps:.1f} fps, {total_frames} frames, {duration:.1f}s")
    
    if frame_step is None:
        frame_step = max(1, int(fps))
    
    print("  Loading EasyOCR...")
    if reader is None:
        reader = load_easyocr()
    
    # ── Step 1: Full OCR on the first few frames to build layout ──────────
    print("  Building layout from first frame...")
    
    # Read several early frames and keep the one with the richest numeric layout.
    candidate_indices = []
    for seconds in (0, 1, 2, 3, 5):
        candidate_idx = int(seconds * fps)
        if candidate_idx < total_frames and candidate_idx not in candidate_indices:
            candidate_indices.append(candidate_idx)

    best_layout = None
    for candidate_idx in candidate_indices:
        cap.set(cv2.CAP_PROP_POS_FRAMES, candidate_idx)
        ret, bgr = cap.read()
        if not ret or bgr is None:
            continue

        candidate_frame = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        text_items_candidate, num_items_candidate = auto_detect_layout(reader, candidate_frame, params)
        score = len(num_items_candidate)
        if best_layout is None or score > best_layout['score']:
            best_layout = {
                'frame_idx': candidate_idx,
                'frame_rgb': candidate_frame,
                'text_items': text_items_candidate,
                'num_items': num_items_candidate,
                'score': score,
            }

        if score >= max(5, min(20, len(params) // 6)):
            break

    if best_layout is None:
        print("  ERROR: Cannot read initial frames for layout detection")
        cap.release()
        return []

    first_frame = best_layout['frame_rgb']
    text_items = best_layout['text_items']
    num_items = best_layout['num_items']
    print(f"  Layout source frame: {best_layout['frame_idx']}")
    
    print(f"  Found {len(num_items)} numeric positions in first frame")
    
    # Build sorted layout: each numeric position becomes one "column" in output
    # Sort by position: top-to-bottom, left-to-right
    num_items_sorted = sorted(num_items, key=lambda x: (int(x[1] / 40) * 10000 + int(x[0] / 40)))
    
    # Prepare "regions" - each region is a bounding box around a number position
    REGION_PAD = 15  # pixels
    regions = []
    for cx, cy, val, text, conf, bbox in num_items_sorted:
        # Get bbox bounds
        pts = np.array(bbox)
        x1 = max(0, int(pts[:, 0].min()) - REGION_PAD)
        y1 = max(0, int(pts[:, 1].min()) - REGION_PAD)
        x2 = min(width, int(pts[:, 0].max()) + REGION_PAD)
        y2 = min(height, int(pts[:, 1].max()) + REGION_PAD)
        regions.append({
            'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2,
            'cx': float(cx), 'cy': float(cy),
            'sample_val': val,
            'label': None,
            'param_name': None,
        })
    
    # Try to match regions to parameter names using nearby text labels
    regions = label_regions_from_text(regions, text_items, params)
    
    print(f"  Layout has {len(regions)} value regions")
    
    # ── Step 2: Process all frames ─────────────────────────────────────────
    print(f"  Processing all frames (step={frame_step})...")
    
    all_rows = []
    frame_idx = 0
    processed = 0
    
    while True:
        if max_frames and processed >= max_frames:
            break
        
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ret, bgr = cap.read()
        if not ret:
            break
        
        timestamp = frame_idx / fps
        frame_rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        frame_pil = Image.fromarray(frame_rgb)
        frame_started_at = time.perf_counter()
        
        row_data = {
            '_frame': frame_idx,
            '_timestamp': f"{int(timestamp//3600):02d}:{int((timestamp%3600)//60):02d}:{int(timestamp%60):02d}",
        }
        
        # For each region, crop and OCR
        for reg_idx, reg in enumerate(regions):
            crop = frame_rgb[reg['y1']:reg['y2'], reg['x1']:reg['x2']]
            if crop.size == 0:
                row_data[f'reg_{reg_idx+1}'] = None
                continue
            
            # Upscale small crops for better OCR
            h, w = crop.shape[:2]
            if h < 20 or w < 20:
                scale = max(20 / h, 20 / w, 1) * 3
                crop = cv2.resize(crop, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
            
            try:
                result = reader.readtext(crop, detail=1, allowlist='0123456789.-+', paragraph=False)
                if result:
                    normalized = [tuple(item) for item in result if isinstance(item, (list, tuple)) and len(item) >= 3]
                    if not normalized:
                        row_data[f'reg_{reg_idx+1}'] = None
                        continue
                    best = max(normalized, key=lambda item: float(item[2]))
                    best_text = str(best[1])
                    val = parse_number(best_text)
                    row_data[f'reg_{reg_idx+1}'] = val if val is not None else best_text
                else:
                    row_data[f'reg_{reg_idx+1}'] = None
            except Exception:
                row_data[f'reg_{reg_idx+1}'] = None

        row_data['_frame_processing_seconds'] = round(time.perf_counter() - frame_started_at, 4)
        
        all_rows.append(row_data)
        processed += 1
        
        if processed % 10 == 0 or processed == 1:
            print(f"    Frame {frame_idx}/{total_frames} (t={timestamp:.1f}s), processed={processed}")
        
        frame_idx += frame_step
    
    cap.release()
    print(f"  Done. Total processed: {processed} frames")
    
    # ── Step 3: Save to Excel ──────────────────────────────────────────────
    output_path = OUTPUT_DIR / f"Результаты_смарт_{video_idx}.xlsx"
    save_smart_results(all_rows, regions, params, output_path, video_idx)
    print(f"  Saved: {output_path}")
    
    return all_rows


def save_smart_results(all_rows, regions, params, output_path, video_idx):
    from openpyxl.styles import PatternFill, Font, Alignment
    import time
    
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = f"Видео {video_idx}"
    
    n_regions = len(regions)
    
    # Build header: frame, timestamp, reg_1...reg_N
    headers = ['Кадр', 'Время', 'Обработка кадра, сек']
    for i, reg in enumerate(regions):
        label = reg.get('label', '')
        param = reg.get('param_name', '')
        sample = reg.get('sample_val', '')
        if param:
            # Use param name (truncated)
            short_param = param[:40] if len(param) > 40 else param
            headers.append(short_param)
        elif label:
            headers.append(f'{label} (~{sample})')
        else:
            headers.append(f'Рег.{i+1} (~{sample})')
    
    ws.append(headers)
    
    # Style header
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 45
    
    # Data
    for row in all_rows:
        data = [
            row['_frame'],
            row['_timestamp'],
            row.get('_frame_processing_seconds', ''),
        ]
        for i in range(1, n_regions + 1):
            data.append(row.get(f'reg_{i}', ''))
        ws.append(data)
    
    # Freeze first row and metadata columns
    ws.freeze_panes = 'D2'
    
    # Regions info sheet
    ws2 = wb.create_sheet("Позиции регионов")
    ws2.append(['#', 'X1', 'Y1', 'X2', 'Y2', 'Центр X', 'Центр Y',
                'Значение (1-й кадр)', 'Метка OCR', 'Параметр таблицы'])
    for i, reg in enumerate(regions, 1):
        ws2.append([i, reg['x1'], reg['y1'], reg['x2'], reg['y2'],
                    int(reg['cx']), int(reg['cy']), reg['sample_val'],
                    reg.get('label', ''), reg.get('param_name', '')])
    
    # Parameters sheet
    ws3 = wb.create_sheet("Параметры")
    ws3.append(['#', 'Наименование', 'Ед.изм.', 'Кор.назв.'])
    for i, p in enumerate(params, 1):
        ws3.append([i, p['name'], p['unit'], p['short']])
    
    # Try to remove existing file to avoid PermissionError; if unable, append timestamp
    try:
        if output_path.exists():
            output_path.unlink()
        wb.save(output_path)
    except PermissionError:
        alt = output_path.with_name(output_path.stem + f"_{int(time.time())}" + output_path.suffix)
        wb.save(alt)


def process_video_full_ocr(video_idx, frame_step=None, max_frames=None, reader=None):
    """
    Full OCR per frame — no pre-built layout.
    Reads ALL numbers from each frame and associates them with table parameters.
    
    Association strategy:
    - The table has parameter names with short labels (T, P, dP, N, etc.)
    - OCR both labels and numbers, find label→nearby number pairs
    - Match found parameters to table by short name + ordinal (001, 002, etc.)
    """
    video_dir = BASE / f"Видео {video_idx}"
    video_path = video_dir / f"Видео {video_idx}.mp4"
    table_path = video_dir / f"Таблица {video_idx}.xlsx"
    
    print(f"\n{'='*60}")
    print(f"[Full OCR] Processing: Видео {video_idx}")
    
    params = load_table_params(table_path)
    print(f"  Table parameters: {len(params)}")
    
    cap = cv2.VideoCapture(str(video_path))
    fps = cap.get(cv2.CAP_PROP_FPS)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    duration = total_frames / fps if fps > 0 else 0
    
    print(f"  {width}x{height} @ {fps:.1f}fps, {total_frames} frames, {duration:.1f}s")
    
    if frame_step is None:
        frame_step = max(1, int(fps))  # 1 sample/second
    
    print("  Loading OCR (ru+en)...")
    if reader is None:
        reader = OCRAdapter('easyocr')
    
    all_rows = []
    frame_idx = 0
    processed = 0
    
    while True:
        if max_frames and processed >= max_frames:
            break
        
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
        ret, bgr = cap.read()
        if not ret:
            break
        
        timestamp = frame_idx / fps
        frame_rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        
        # Run full OCR
        ocr_results = ocr_frame(reader, frame_rgb)
        
        # Build maps
        all_numbers = []
        all_words = []
        
        for bbox, text, conf in ocr_results:
            cx, cy = bbox_center(bbox)
            val = parse_number(text)
            conf_value = float(conf)
            if val is not None and conf_value > 0.25:
                all_numbers.append({
                    'cx': float(cx), 'cy': float(cy),
                    'val': val, 'text': text, 'conf': conf_value, 'bbox': bbox
                })
            elif conf_value > 0.35 and len(text.strip()) > 0:
                all_words.append({
                    'cx': float(cx), 'cy': float(cy),
                    'text': text.strip(), 'conf': conf_value, 'bbox': bbox
                })
        
        row_data = {
            '_frame': frame_idx,
            '_timestamp': f"{int(timestamp//3600):02d}:{int((timestamp%3600)//60):02d}:{int(timestamp%60):02d}",
            '_n_numbers': len(all_numbers),
            '_n_words': len(all_words),
        }
        
        # Store raw numbers sorted by screen position
        nums_sorted = sorted(all_numbers, key=lambda x: (int(x['cy'] / 40) * 100000 + int(x['cx'] / 10)))
        for i, n in enumerate(nums_sorted):
            row_data[f'n{i+1}_val'] = n['val']
            row_data[f'n{i+1}_x'] = int(n['cx'])
            row_data[f'n{i+1}_y'] = int(n['cy'])
        
        all_rows.append(row_data)
        processed += 1
        
        if processed % 5 == 0 or processed == 1:
            print(f"    [{processed}] frame={frame_idx}/{total_frames} t={timestamp:.1f}s | {len(all_numbers)} nums, {len(all_words)} words")
        
        frame_idx += frame_step
    
    cap.release()
    print(f"  Processed {processed} frames total")
    
    output_path = OUTPUT_DIR / f"Результаты_fullOCR_{video_idx}.xlsx"
    save_full_ocr_results(all_rows, params, output_path, video_idx)
    print(f"  Saved: {output_path}")
    
    return all_rows


def save_full_ocr_results(all_rows, params, output_path, video_idx):
    from openpyxl.styles import PatternFill, Font, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Failed to create worksheet")
    ws.title = f"Числа из кадров"
    
    # Dynamic columns based on max numbers found
    max_nums = max((row.get('_n_numbers', 0) for row in all_rows), default=0)
    
    headers = ['Кадр', 'Время', 'Кол-во чисел', 'Кол-во слов']
    for i in range(1, max_nums + 1):
        headers.extend([f'N{i}', f'X{i}', f'Y{i}'])
    
    ws.append(headers)
    
    fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    for row in all_rows:
        data = [row['_frame'], row['_timestamp'], row.get('_n_numbers', 0), row.get('_n_words', 0)]
        for i in range(1, max_nums + 1):
            data.extend([
                row.get(f'n{i}_val', ''),
                row.get(f'n{i}_x', ''),
                row.get(f'n{i}_y', ''),
            ])
        ws.append(data)
    
    # Params sheet
    ws2 = wb.create_sheet("Параметры таблицы")
    ws2.append(['#', 'Наименование', 'Ед.изм.', 'Кор.назв.', 'Знаков'])
    for i, p in enumerate(params, 1):
        ws2.append([i, p['name'], p['unit'], p['short'], p['decimals']])
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="SCADA Mnemonic Video → Excel Extractor")
    parser.add_argument('--video_idx', type=int, choices=[1, 2, 3], help='Video index (1, 2, or 3)')
    parser.add_argument('--all', action='store_true', help='Process all three videos')
    parser.add_argument('--mode', choices=['smart', 'full', 'both'], default='smart',
                        help='Extraction mode: smart (region-based) or full (full-OCR per frame)')
    parser.add_argument('--frame_step', type=int, default=None,
                        help='Frame step (default: 1 per second based on FPS)')
    parser.add_argument('--max_frames', type=int, default=None,
                        help='Max frames to process (for testing)')
    
    args = parser.parse_args()
    
    indices = []
    if args.all:
        indices = [1, 2, 3]
    elif args.video_idx:
        indices = [args.video_idx]
    else:
        print("Specify --video_idx 1/2/3 or --all")
        parser.print_help()
        return
    
    for idx in indices:
        if args.mode in ('smart', 'both'):
            process_video_smart(idx, frame_step=args.frame_step, max_frames=args.max_frames)
        if args.mode in ('full', 'both'):
            process_video_full_ocr(idx, frame_step=args.frame_step, max_frames=args.max_frames)


if __name__ == '__main__':
    main()
