"""Compare Paddle (paddlevl) and EasyOCR on Видео 3 with per-frame noise detection.

Writes per-frame Excel with columns: frame_idx, easyocr_count, paddle_count, lap_var, noise_std, diff
Also writes aggregate comparison sheet.

Usage (from repo root, with venv):
  $env:PADDLE_DISABLE_ONEDNN='1'; $env:FLAGS_use_mkldnn='0'; \
  .\.venv\Scripts\python.exe scripts\compare_paddle_easy_noise.py --max-frames 200

"""
from pathlib import Path
import sys
import time
import argparse

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import cv2
import numpy as np
import openpyxl

from extract_scada import process_video_smart, OCRAdapter

VIDEO_DIR = ROOT / 'Хакатон «ИИ – АВТОМАТИЗАЦИЯ»' / 'Задание 2' / 'Видео 3'
VIDEO_FILE = VIDEO_DIR / 'Видео 3.mp4'
TABLE_PATH = VIDEO_DIR / 'Таблица 3.xlsx'
OUTDIR = ROOT / 'results'
OUTDIR.mkdir(exist_ok=True)

parser = argparse.ArgumentParser()
parser.add_argument('--max-frames', type=int, default=200)
parser.add_argument('--frame-step', type=int, default=1)
parser.add_argument('--paddle-preprocess', type=str, default='none', help='preprocess to apply for paddlevl: none|clahe|hist|sharpen|thres')
args = parser.parse_args()
MAX_FRAMES = args.max_frames
FRAME_STEP = args.frame_step
PADDLE_PRE = args.paddle_preprocess

def preprocess_none(img):
    return img

def preprocess_clahe(img):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    g2 = clahe.apply(g)
    return cv2.cvtColor(g2, cv2.COLOR_GRAY2BGR)

def preprocess_hist(img):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    g2 = cv2.equalizeHist(g)
    return cv2.cvtColor(g2, cv2.COLOR_GRAY2BGR)

def preprocess_sharpen(img):
    kernel = np.array([[0,-1,0],[-1,5,-1],[0,-1,0]])
    return cv2.filter2D(img, -1, kernel)

def preprocess_thres(img):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, t = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    return cv2.cvtColor(t, cv2.COLOR_GRAY2BGR)

PRE_MAP = {
    'none': preprocess_none,
    'clahe': preprocess_clahe,
    'hist': preprocess_hist,
    'sharpen': preprocess_sharpen,
    'thres': preprocess_thres,
}

def make_reader(name):
    if name == 'easyocr':
        return OCRAdapter('easyocr')
    if name == 'paddlevl':
        base = OCRAdapter('paddlevl')
        pre = PRE_MAP.get(PADDLE_PRE, preprocess_none)
        # wrapper object with readtext method
        class Wrapped:
            def readtext(self, img, *a, **k):
                return base.readtext(pre(img), *a, **k)
        return Wrapped()
    return OCRAdapter(name)

variants = [
    ('easyocr', lambda: make_reader('easyocr')),
    ('paddlevl', lambda: make_reader('paddlevl')),
]

# Run OCR variants and collect rows per-frame
all_rows = {}
all_stats = []
for name, ctor in variants:
    print(f"\n--- Running {name} ---")
    try:
        reader = ctor()
    except Exception as e:
        print('Init error for', name, e)
        all_rows[name] = []
        all_stats.append({'model': name, 'frames': 0, 'time_s': 0.0, 'error': str(e)})
        continue
    start = time.time()
    rows = process_video_smart(3, frame_step=FRAME_STEP, max_frames=MAX_FRAMES, reader=reader)
    elapsed = time.time() - start
    print(f'Finished {name}: frames={len(rows)} time={elapsed:.2f}s')
    all_rows[name] = rows
    non_empty = 0
    for r in rows:
        for k, v in r.items():
            if k.startswith('reg_') and v not in (None, ''):
                non_empty += 1
    avg = non_empty / len(rows) if rows else 0
    all_stats.append({'model': name, 'frames': len(rows), 'time_s': round(elapsed,2), 'avg_values_per_frame': round(avg,3), 'error': ''})

# Read video frames (first N frames according to processed frames of first variant)
first_variant = variants[0][0]
N = max(len(all_rows.get(first_variant, [])), 1)
N = min(N, MAX_FRAMES)
cap = cv2.VideoCapture(str(VIDEO_FILE))
frames = []
count = 0
while count < N:
    ret, f = cap.read()
    if not ret:
        break
    if FRAME_STEP > 1:
        # skip frames according to step
        pass
    frames.append(f)
    count += 1
cap.release()

# Helper noise metrics
def laplacian_var(img):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return float(cv2.Laplacian(g, cv2.CV_64F).var())

def noise_std_estimate(img):
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY).astype(np.float32)
    blur = cv2.GaussianBlur(g, (3,3), 0)
    noise = g - blur
    return float(np.std(noise))

# Build per-frame table
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'per_frame'
ws.append(['frame_idx','easyocr_count','paddlevl_count','lap_var','noise_std','count_diff'])

M = max(len(all_rows.get('easyocr', [])), len(all_rows.get('paddlevl', [])))
M = min(M, len(frames))
for i in range(M):
    e_rows = all_rows.get('easyocr', [])
    p_rows = all_rows.get('paddlevl', [])
    e_count = 0
    p_count = 0
    if i < len(e_rows):
        for k,v in e_rows[i].items():
            if k.startswith('reg_') and v not in (None, ''):
                e_count += 1
    if i < len(p_rows):
        for k,v in p_rows[i].items():
            if k.startswith('reg_') and v not in (None, ''):
                p_count += 1
    lap = laplacian_var(frames[i]) if i < len(frames) else None
    nstd = noise_std_estimate(frames[i]) if i < len(frames) else None
    diff = e_count - p_count
    ws.append([i, e_count, p_count, lap, nstd, diff])

out_file = OUTDIR / 'per_frame_paddle_easy_noise.xlsx'
wb.save(out_file)
print('Saved per-frame results to', out_file)

# Aggregate comparison
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'summary'
ws2.append(['model','frames','time_s','avg_values_per_frame','error'])
for s in all_stats:
    ws2.append([s.get('model'), s.get('frames'), s.get('time_s'), s.get('avg_values_per_frame'), s.get('error')])

# Simple aggregate comparisons
# compute avg lap and avg noise for frames
import statistics
lap_vals = []
noise_vals = []
wb3 = openpyxl.load_workbook(out_file)
ws3 = wb3['per_frame']
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row[3] is not None:
        lap_vals.append(row[3])
    if row[4] is not None:
        noise_vals.append(row[4])
avg_lap = statistics.mean(lap_vals) if lap_vals else 0
avg_noise = statistics.mean(noise_vals) if noise_vals else 0
ws2.append([])
ws2.append(['avg_laplacian_variance', avg_lap])
ws2.append(['avg_noise_std', avg_noise])

cmp_file = OUTDIR / 'comparison_paddle_easy_noise.xlsx'
wb2.save(cmp_file)
print('Saved comparison summary to', cmp_file)

# Also insert per-frame sheet into Таблица 3.xlsx
try:
    tb = openpyxl.load_workbook(TABLE_PATH)
    name = 'Paddle_vs_Easy_per_frame'[:31]
    if name in tb.sheetnames:
        del tb[name]
    ws_new = tb.create_sheet(name)
    # write from per-frame workbook
    for r in ws.iter_rows(values_only=True):
        ws_new.append(r)
    tb.save(TABLE_PATH)
    print('Inserted per-frame sheet into', TABLE_PATH)
except Exception as e:
    print('Could not modify Таблица 3.xlsx:', e)

print('Done')
