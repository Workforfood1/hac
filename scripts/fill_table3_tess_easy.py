"""Run Video 3 with EasyOCR and Tesseract (if available), write sheets into Таблица 3.xlsx and comparison.

Usage (venv):
  $env:FLAGS_use_mkldnn='0'; $env:PADDLE_DISABLE_ONEDNN='1'; .\.venv\Scripts\python.exe scripts\fill_table3_tess_easy.py
"""
from pathlib import Path
import sys
import time
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from extract_scada import process_video_smart, OCRAdapter

BASE = ROOT / 'Хакатон «ИИ – АВТОМАТИЗАЦИЯ»' / 'Задание 2' / 'Видео 3'
TABLE_PATH = BASE / 'Таблица 3.xlsx'
OUTDIR = ROOT / 'results'
OUTDIR.mkdir(exist_ok=True)

variants = [
    ('easyocr', lambda: OCRAdapter('easyocr')),
    ('tesseract', lambda: OCRAdapter('tesseract')),
]

results = []
for name, ctor in variants:
    print(f"\n=== Running {name} ===")
    try:
        reader = ctor()
    except Exception as e:
        print('Init error for', name, e)
        results.append({'model': name, 'frames': 0, 'time_s': 0.0, 'avg_values_per_frame': 0.0, 'error': str(e)})
        continue
    try:
        start = time.time()
        rows = process_video_smart(3, frame_step=None, max_frames=None, reader=reader)
        elapsed = time.time() - start
        frames = len(rows)
        non_empty = 0
        for r in rows:
            for k, v in r.items():
                if k.startswith('reg_') and v not in (None, ''):
                    non_empty += 1
        avg = non_empty / frames if frames else 0
        results.append({'model': name, 'frames': frames, 'time_s': round(elapsed,2), 'avg_values_per_frame': round(avg,3), 'error': ''})
        # save model-specific results
        out_model = OUTDIR / f'Результаты_смарт_3_{name}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        # collect keys
        if rows:
            keys = []
            for r in rows:
                for k in r.keys():
                    if k not in keys:
                        keys.append(k)
            ws.append(keys)
            for r in rows:
                ws.append([r.get(k, '') for k in keys])
        else:
            ws.append(['No data'])
        wb.save(out_model)
        print('Saved', out_model)
        # insert sheet into original table
        wb2 = openpyxl.load_workbook(TABLE_PATH)
        sheet_name = f'Результаты_{name}'[:31]
        if sheet_name in wb2.sheetnames:
            del wb2[sheet_name]
        ws2 = wb2.create_sheet(sheet_name)
        if rows:
            ws2.append(keys)
            for r in rows:
                ws2.append([r.get(k, '') for k in keys])
        else:
            ws2.append(['No data'])
        wb2.save(TABLE_PATH)
        print('Inserted sheet into Таблица 3.xlsx as', sheet_name)
    except Exception as e:
        print('Run error for', name, e)
        results.append({'model': name, 'frames': 0, 'time_s': 0.0, 'avg_values_per_frame': 0.0, 'error': str(e)})

# write comparison
out_cmp = OUTDIR / 'compare_tesseract_easyocr.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['model','frames','time_s','avg_values_per_frame','error'])
for r in results:
    ws.append([r['model'], r['frames'], r['time_s'], r['avg_values_per_frame'], r['error']])
wb.save(out_cmp)
print('Saved comparison to', out_cmp)
